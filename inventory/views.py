from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.core.paginator import Paginator
from django.http import JsonResponse
from datetime import datetime, timedelta
from decimal import Decimal

from .models import (OfficeSupply, StockInApplication, StockInItem, StockOutRecord, 
                    StockOutOrder, StockOutItem, ITDevice, ComputerType, ReturnApplication, 
                    ItemCategory, Department, Profile, SystemRole,
                    get_role_choices, get_role_group_map, get_role_display_name)
from .forms import (OfficeSupplyForm, StockInApplicationForm, StockInApprovalForm,
                    StockOutForm, ReturnApplicationForm, ComputerTypeForm, ITDeviceForm, 
                    ExcelImportForm, ItemCategoryForm, DepartmentForm,
                    RegisterForm, UserEditForm, PasswordResetForm, ProfileForm)
from .decorators import role_required
from .utils import (get_user_role, get_visible_queryset, check_dept_head_exists, 
                    check_pending_before_role_change)

import json

from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO


# ==================== 部门管理（树型结构）====================
@login_required
def department_list(request):
    """部门一览（树型结构）"""
    departments = Department.objects.filter(parent__isnull=True).order_by('sort_order', 'code')
    return render(request, 'inventory/department_list.html', {'departments': departments})


@login_required
def department_create(request):
    """部门追加"""
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '部门添加成功！')
            return redirect('department_list')
    else:
        initial = {}
        parent_id = request.GET.get('parent')
        if parent_id:
            try:
                parent = Department.objects.get(pk=parent_id)
                initial['parent'] = parent
            except Department.DoesNotExist:
                pass
        form = DepartmentForm(initial=initial)

    return render(request, 'inventory/department_form.html', {
        'form': form,
        'title': '部门追加',
        'action': '创建'
    })


@login_required
def department_update(request, pk):
    """部门变更"""
    department = get_object_or_404(Department, pk=pk)
    if request.method == 'POST':
        form = DepartmentForm(request.POST, instance=department)
        if form.is_valid():
            form.save()
            messages.success(request, '部门更新成功！')
            return redirect('department_list')
    else:
        form = DepartmentForm(instance=department)

    return render(request, 'inventory/department_form.html', {
        'form': form,
        'title': '部门变更',
        'action': '更新'
    })


@login_required
def department_delete(request, pk):
    """部门删除"""
    department = get_object_or_404(Department, pk=pk)
    if request.method == 'POST':
        if department.children.exists():
            messages.error(request, '该部门下存在子部门，无法删除！')
            return redirect('department_list')
        department.delete()
        messages.success(request, '部门删除成功！')
        return redirect('department_list')

    return render(request, 'inventory/department_confirm_delete.html', {'department': department})


# ==================== 首页 ====================
@login_required
def home(request):
    """系统首页"""
    role = get_user_role(request.user)
    
    # 统计数据（按角色过滤）
    supply_count = OfficeSupply.objects.count()
    low_stock_count = OfficeSupply.objects.filter(status='低库存').count()
    visible_stockin = get_visible_queryset(request.user, StockInApplication)
    visible_stockout = get_visible_queryset(request.user, StockOutOrder)
    pending_count = visible_stockin.filter(status='待审批').count()
    pending_stockout_count = visible_stockout.filter(status='待审批').count()
    device_count = get_visible_queryset(request.user, ITDevice).count()
    
    # 最近入库单
    recent_applications = visible_stockin.select_related('applicant', 'department').prefetch_related('items__supply').order_by('-created_at')[:5]
    
    # 最近出库记录
    recent_outs = visible_stockout.prefetch_related('items__supply').order_by('-created_at')[:5]
    
    context = {
        'supply_count': supply_count,
        'low_stock_count': low_stock_count,
        'pending_count': pending_count,
        'pending_stockout_count': pending_stockout_count,
        'device_count': device_count,
        'recent_applications': recent_applications,
        'recent_outs': recent_outs,
    }
    return render(request, 'inventory/home.html', context)


# ==================== 物品分类管理（树型结构）====================
@login_required
def item_category_list(request):
    """物品分类一览（树型结构）"""
    categories = ItemCategory.objects.filter(parent__isnull=True).order_by('sort_order', 'code')
    return render(request, 'inventory/item_category_list.html', {'categories': categories})


@login_required
def item_category_create(request):
    """分类追加"""
    if request.method == 'POST':
        form = ItemCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '分类添加成功！')
            return redirect('item_category_list')
    else:
        form = ItemCategoryForm()
    
    return render(request, 'inventory/item_category_form.html', {
        'form': form,
        'title': '分类追加',
        'action': '创建'
    })


@login_required
def item_category_update(request, pk):
    """分类变更"""
    category = get_object_or_404(ItemCategory, pk=pk)
    if request.method == 'POST':
        form = ItemCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, '分类更新成功！')
            return redirect('item_category_list')
    else:
        form = ItemCategoryForm(instance=category)
    
    return render(request, 'inventory/item_category_form.html', {
        'form': form,
        'title': '分类变更',
        'action': '更新'
    })


@login_required
def item_category_delete(request, pk):
    """分类删除"""
    category = get_object_or_404(ItemCategory, pk=pk)
    if request.method == 'POST':
        if category.children.exists():
            messages.error(request, '该分类下存在子分类，无法删除！')
            return redirect('item_category_list')
        category.delete()
        messages.success(request, '分类删除成功！')
        return redirect('item_category_list')
    
    return render(request, 'inventory/item_category_confirm_delete.html', {'category': category})


# ==================== 办公用品库存管理 ====================
@login_required
def supply_list(request):
    """库存一览"""
    query = request.GET.get('q', '')
    category_id = request.GET.get('category', '')
    status = request.GET.get('status', '')
    
    supplies = OfficeSupply.objects.select_related('item_category').all()
    
    if query:
        supplies = supplies.filter(
            Q(code__icontains=query) | Q(name__icontains=query)
        )
    if category_id:
        supplies = supplies.filter(item_category_id=category_id)
    if status:
        supplies = supplies.filter(status=status)
    
    # 分页
    paginator = Paginator(supplies, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # 获取类别列表用于筛选
    categories = ItemCategory.objects.filter(is_active=True).order_by('sort_order', 'code')
    
    context = {
        'page_obj': page_obj,
        'categories': categories,
        'query': query,
        'category_filter': category_id,
        'status_filter': status,
    }
    return render(request, 'inventory/supply_list.html', context)


@login_required
def supply_name_search(request):
    """API: 按拼音首字母/编码/名称模糊匹配返回已有物品列表"""
    try:
        from pypinyin import lazy_pinyin
    except Exception:
        lazy_pinyin = None

    q = request.GET.get('q', '').strip()
    if not q:
        return JsonResponse({'results': []})

    # 默认过滤停用物品（包含历史脏数据：如“停用 ”、“已停用”等）
    include_disabled = request.GET.get('include_disabled') == '1'

    q_upper = q.upper()
    results = []
    supplies_qs = OfficeSupply.objects.all()
    if not include_disabled:
        supplies_qs = supplies_qs.exclude(
            Q(status__icontains='停用') | Q(item_category__is_active=False)
        )

    supplies = supplies_qs.select_related('item_category').order_by('code')
    for s in supplies:
        py = ''
        if lazy_pinyin:
            py = ''.join([p[0].upper() for p in lazy_pinyin(s.name) if p])

        if (
            q_upper in s.name.upper()
            or q_upper in s.code.upper()
            or (py and q_upper in py)
        ):
            results.append({
                'id': s.id,
                'code': s.code,
                'name': s.name,
                'specification': s.specification or '',
                'item_category_id': s.item_category_id,
                'item_category_name': s.item_category.name if s.item_category else '',
                'pinyin': py,
                'quantity': s.quantity,
                'available_quantity': s.available_quantity,
                'unit': s.unit or '',
                'price': str(s.price),
            })
        if len(results) >= 20:
            break
    return JsonResponse({'results': results})


def supply_create(request):
    """物品信息登记 - 仅创建物品主数据，不设置库存（库存通过入库单增加）"""
    if request.method == 'POST':
        form = OfficeSupplyForm(request.POST)
        if form.is_valid():
            supply = form.save(commit=False)
            supply.quantity = 0
            supply.status = '正常'
            supply.save()
            # 模型 save() 会自动根据 quantity<=safety_stock 设状态，新物品 quantity=0 会被改为「低库存」，
            # 此处覆盖回「正常」——库存为零是预期状态，实际库存只通过入库单增加
            if supply.status == '低库存':
                supply.status = '正常'
                supply.save(update_fields=['status'])
            messages.success(request, f'物品 "{supply.name}" 登记成功！库存请通过入库单增加。')
            return redirect('supply_list')
    else:
        form = OfficeSupplyForm()
    
    return render(request, 'inventory/supply_form.html', {
        'form': form,
        'title': '物品信息登记',
        'action': '创建'
    })


@login_required
def supply_update(request, pk):
    """库存信息变更"""
    supply = get_object_or_404(OfficeSupply, pk=pk)
    if request.method == 'POST':
        form = OfficeSupplyForm(request.POST, instance=supply)
        if form.is_valid():
            form.save()
            messages.success(request, '库存信息更新成功！')
            return redirect('supply_list')
    else:
        form = OfficeSupplyForm(instance=supply)
    
    return render(request, 'inventory/supply_form.html', {
        'form': form,
        'supply': supply,
        'title': '物品信息变更',
        'action': '更新'
    })


@login_required
def supply_delete(request, pk):
    """删除库存记录"""
    supply = get_object_or_404(OfficeSupply, pk=pk)
    if request.method == 'POST':
        supply.delete()
        messages.success(request, '库存记录已删除！')
        return redirect('supply_list')
    return render(request, 'inventory/supply_confirm_delete.html', {'supply': supply})


@login_required
def supply_add_stock(request, pk):
    """库存追加"""
    supply = get_object_or_404(OfficeSupply, pk=pk)
    if request.method == 'POST':
        add_qty = int(request.POST.get('quantity', 0))
        if add_qty > 0:
            supply.quantity += add_qty
            supply.save()
            messages.success(request, f'已为 "{supply.name}" 追加 {add_qty} {supply.unit}')
        return redirect('supply_list')
    return render(request, 'inventory/supply_add_stock.html', {'supply': supply})


# ==================== 入库单管理 ====================
@login_required
def stockin_application_list(request):
    """入库单查询（支持多物品）"""
    query = request.GET.get('q', '')
    status = request.GET.get('status', '')

    applications = get_visible_queryset(request.user, StockInApplication).prefetch_related('items__supply').select_related('applicant')

    if query:
        applications = applications.filter(
            Q(application_no__icontains=query) | Q(items__supply__name__icontains=query)
        ).distinct()
    if status:
        applications = applications.filter(status=status)

    paginator = Paginator(applications, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'query': query,
        'status_filter': status,
    }
    return render(request, 'inventory/stockin_application_list.html', context)


@login_required
def stockin_application_create(request):
    """入库单登记（支持多物品，只选已有）"""
    if request.method == 'POST':
        form = StockInApplicationForm(request.POST)
        items_json = request.POST.get('items_json', '[]')
        import json
        try:
            items_data = json.loads(items_json)
        except (json.JSONDecodeError, TypeError):
            items_data = []

        if form.is_valid() and items_data:
            # 校验物品：必须全部存在且为启用状态
            errors = []
            for item in items_data:
                supply_id = item.get('supply_id')
                qty = int(item.get('quantity', 0))
                try:
                    supply = OfficeSupply.objects.select_related('item_category').get(pk=supply_id)
                except OfficeSupply.DoesNotExist:
                    errors.append(f'物品 ID {supply_id} 不存在')
                    continue
                is_supply_disabled = ('停用' in (supply.status or '')) or (
                    supply.item_category is not None and not supply.item_category.is_active
                )
                if is_supply_disabled:
                    errors.append(f'{supply.code} - {supply.name} 已停用，禁止入库')
                elif qty <= 0:
                    errors.append(f'{supply.name}：入库数量必须大于 0')

            if errors:
                for err in errors:
                    messages.error(request, err)
                return render(request, 'inventory/stockin_application_form.html', {
                    'form': form, 'title': '入库单', 'action': '提交'
                })

            # 创建申请单
            app = form.save(commit=False)
            app.applicant = request.user
            app.save()

            # 创建明细（保留每行独立记录）
            for item in items_data:
                supply = OfficeSupply.objects.get(pk=item['supply_id'])
                qty = int(item['quantity'])
                unit_price = Decimal(str(item.get('unit_price', supply.price or 0)))
                StockInItem.objects.create(
                    application=app, supply=supply, quantity=qty, unit_price=unit_price,
                    specification=supply.specification or '',
                    unit=supply.unit,
                    location=supply.location or '',
                    supplier=supply.supplier or '',
                    doc_no=item.get('doc_no', ''),
                )

            messages.success(request, f'入库单 "{app.application_no}" 提交成功，共 {len(items_data)} 项物品！')
            return redirect('stockin_application_list')
        else:
            if not items_data:
                messages.error(request, '请至少添加一项入库物品')
    else:
        form = StockInApplicationForm()

    return render(request, 'inventory/stockin_application_form.html', {
        'form': form,
        'title': '入库单',
        'action': '提交',
        'existing_items_json': '[]',
    })


@login_required
def stockin_application_update(request, pk):
    """入库单变更（支持多物品）"""
    application = get_object_or_404(StockInApplication, pk=pk)
    if application.status != '待审批':
        messages.error(request, '已审批的申请不能修改！')
        return redirect('stockin_application_list')

    if request.method == 'POST':
        form = StockInApplicationForm(request.POST, instance=application)
        items_json = request.POST.get('items_json', '[]')
        import json
        try:
            items_data = json.loads(items_json)
        except (json.JSONDecodeError, TypeError):
            items_data = []

        if form.is_valid() and items_data:
            # 校验物品：必须全部存在且为启用状态
            errors = []
            for item in items_data:
                supply_id = item.get('supply_id')
                qty = int(item.get('quantity', 0))
                try:
                    supply = OfficeSupply.objects.select_related('item_category').get(pk=supply_id)
                except OfficeSupply.DoesNotExist:
                    errors.append(f'物品 ID {supply_id} 不存在')
                    continue
                is_supply_disabled = ('停用' in (supply.status or '')) or (
                    supply.item_category is not None and not supply.item_category.is_active
                )
                if is_supply_disabled:
                    errors.append(f'{supply.code} - {supply.name} 已停用，禁止入库')
                elif qty <= 0:
                    errors.append(f'{supply.name}：入库数量必须大于 0')

            if errors:
                for err in errors:
                    messages.error(request, err)
                return render(request, 'inventory/stockin_application_form.html', {
                    'form': form, 'application': application,
                    'title': '申请变更', 'action': '更新'
                })

            form.save()
            # 重建明细
            application.items.all().delete()
            for item in items_data:
                supply = OfficeSupply.objects.get(pk=item['supply_id'])
                qty = int(item['quantity'])
                unit_price = Decimal(str(item.get('unit_price', supply.price or 0)))
                StockInItem.objects.create(
                    application=application, supply=supply, quantity=qty, unit_price=unit_price,
                    specification=supply.specification or '',
                    unit=supply.unit,
                    location=supply.location or '',
                    supplier=supply.supplier or '',
                    doc_no=item.get('doc_no', ''),
                )

            messages.success(request, '入库单更新成功！')
            return redirect('stockin_application_list')
        else:
            if not items_data:
                messages.error(request, '请至少添加一项入库物品')
    else:
        form = StockInApplicationForm(instance=application)

    # 编辑模式：回显已有明细
    import json
    existing_items = []
    if application:
        for item in application.items.select_related('supply').all():
            existing_items.append({
                'supply_id': item.supply_id,
                'supply_name': item.supply.name + (f' ({item.supply.specification})' if item.supply.specification else ''),
                'specification': item.specification or item.supply.specification or '',
                'quantity': item.quantity,
                'unit_price': float(item.unit_price),
                'doc_no': item.doc_no or '',
            })

    return render(request, 'inventory/stockin_application_form.html', {
        'form': form,
        'application': application,
        'title': '申请变更',
        'action': '更新',
        'existing_items_json': json.dumps(existing_items),
    })


@login_required
def stockin_application_delete(request, pk):
    """入库单删除"""
    application = get_object_or_404(StockInApplication, pk=pk)
    if application.status != '待审批':
        messages.error(request, '已审批的申请不能删除！')
        return redirect('stockin_application_list')

    if request.method == 'POST':
        application.delete()
        messages.success(request, '入库单已删除！')
        return redirect('stockin_application_list')
    return render(request, 'inventory/stockin_application_confirm_delete.html', {'application': application})


@login_required
def stockin_application_detail(request, pk):
    """入库单明细查看"""
    application = get_object_or_404(
        StockInApplication.objects.select_related('applicant', 'department', 'approver').prefetch_related('items__supply'),
        pk=pk
    )
    return render(request, 'inventory/stockin_application_detail.html', {
        'application': application,
    })


# ==================== 审批管理 ====================
@login_required
def approval_list(request):
    """入出库审批（入库+出库合并列表）"""
    status_filter = request.GET.get('status', '待审批')
    type_filter = request.GET.get('type', '')
    role = get_user_role(request.user)

    # 入库单（按角色过滤）
    stockin_qs = get_visible_queryset(request.user, StockInApplication).prefetch_related('items__supply').select_related('applicant', 'department')
    if status_filter:
        stockin_qs = stockin_qs.filter(status=status_filter)
    stockin_list = [{'obj': a, 'type': '入库', 'no': a.application_no, 'summary': a.get_items_summary(),
                     'amount': a.get_total_amount(), 'applicant': a.applicant.username,
                     'department': str(a.department) if a.department else '-', 'status': a.status,
                     'created_at': a.created_at, 'pk': a.pk} for a in stockin_qs]

    # 出库单（按角色过滤）
    stockout_qs = get_visible_queryset(request.user, StockOutOrder).prefetch_related('items__supply').select_related('operator', 'department')
    if status_filter:
        stockout_qs = stockout_qs.filter(status=status_filter)
    stockout_list = [{'obj': o, 'type': '出库', 'no': o.record_no, 'summary': o.get_items_summary(),
                      'amount': None, 'applicant': o.operator.username,
                      'department': str(o.department) if o.department else '-', 'status': o.status,
                      'created_at': o.created_at, 'pk': o.pk} for o in stockout_qs]

    # 合并并按创建时间倒序
    combined = stockin_list + stockout_list
    if type_filter:
        combined = [r for r in combined if r['type'] == type_filter]
    combined.sort(key=lambda x: x['created_at'], reverse=True)

    paginator = Paginator(combined, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'inventory/approval_list.html', {
        'page_obj': page_obj,
        'status_filter': status_filter,
        'type_filter': type_filter,
    })


@login_required
def approval_process(request, pk):
    """审批处理（支持多物品入库）——基于权限配置"""
    from inventory.permissions import has_permission
    role = get_user_role(request.user)
    if not has_permission(role, 'stockin', 'approve'):
        messages.error(request, '您没有入库审批权限')
        return redirect('approval_list')
    
    application = get_object_or_404(StockInApplication, pk=pk)
    if application.status != '待审批':
        messages.error(request, '该入库单已审批')
        return redirect('approval_list')

    if request.method == 'POST':
        action = request.POST.get('action')
        comment = request.POST.get('approval_comment', '')

        if action == 'approve':
            application.status = '已批准'
            # 批准后逐项增加库存
            for item in application.items.select_related('supply').all():
                item.supply.quantity += item.quantity
                item.supply.save()
            messages.success(request, f'申请 "{application.application_no}" 已批准，共 {application.items.count()} 项物品库存已更新！')
        elif action == 'reject':
            application.status = '已拒绝'
            messages.info(request, f'申请 "{application.application_no}" 已拒绝。')

        application.approver = request.user
        application.approval_time = timezone.now()
        application.approval_comment = comment
        application.save()

        return redirect('approval_list')

    return render(request, 'inventory/approval_process.html', {'application': application})


@login_required
def approval_process_stockout(request, pk):
    """二级审批处理（出库单）"""
    from django.db.models import F
    order = get_object_or_404(
        StockOutOrder.objects.select_related('department', 'operator').prefetch_related('items__supply'),
        pk=pk
    )
    role = get_user_role(request.user)

    if request.method == 'POST':
        action = request.POST.get('action')
        comment = request.POST.get('comment', '')

        # 部门长一级审批
        if action == 'dept_approve':
            if order.status != '待审批':
                messages.error(request, '该出库单状态不符，无法审批')
                return redirect('stockout_detail', pk=order.pk)
            from inventory.permissions import get_data_scope
            approve_scope = get_data_scope(role, 'stockout', 'approve')
            if approve_scope not in ('all', 'dept'):
                messages.error(request, '您没有部门长审批权限')
                return redirect('stockout_detail', pk=order.pk)
            order.status = '待仓管审批'
            order.dept_approver = request.user
            order.dept_approval_time = timezone.now()
            order.dept_approval_comment = comment
            order.save()
            messages.success(request, f'出库单 "{order.record_no}" 部门长已通过，待仓管审批')

        elif action == 'dept_reject':
            if order.status != '待审批':
                messages.error(request, '该出库单状态不符')
                return redirect('stockout_detail', pk=order.pk)
            from inventory.permissions import get_data_scope
            approve_scope = get_data_scope(role, 'stockout', 'approve')
            if approve_scope not in ('all', 'dept'):
                messages.error(request, '您没有部门长审批权限')
                return redirect('stockout_detail', pk=order.pk)
            # 驳回时释放锁定库存
            for item in order.items.select_related('supply').all():
                OfficeSupply.objects.filter(pk=item.supply.pk).update(
                    locked_quantity=F('locked_quantity') - item.quantity
                )
            order.status = '已拒绝'
            order.dept_approver = request.user
            order.dept_approval_time = timezone.now()
            order.dept_approval_comment = comment
            order.save()
            messages.info(request, f'出库单 "{order.record_no}" 已被部门长驳回，已释放锁定库存')

        # 仓管二级审批
        elif action == 'wh_approve':
            if order.status != '待仓管审批':
                messages.error(request, '该出库单状态不符，无法审批')
                return redirect('stockout_detail', pk=order.pk)
            from inventory.permissions import has_permission
            if not has_permission(role, 'stockout', 'approve'):
                messages.error(request, '您没有仓管审批权限')
                return redirect('stockout_detail', pk=order.pk)
            # 校验库存
            errors = []
            for item in order.items.select_related('supply').all():
                if item.quantity > item.supply.quantity:
                    errors.append(f'{item.supply.name} 库存不足！库存：{item.supply.quantity}，需要：{item.quantity}')
            if errors:
                for err in errors:
                    messages.error(request, err)
                return redirect('stockout_detail', pk=order.pk)

            # 审批通过：扣减实际库存并释放锁定（原子操作，使用update避免触发save中的F表达式比较）
            for item in order.items.select_related('supply').all():
                OfficeSupply.objects.filter(pk=item.supply.pk).update(
                    quantity=F('quantity') - item.quantity,
                    locked_quantity=F('locked_quantity') - item.quantity
                )
            order.status = '已批准'
            order.approver = request.user
            order.approval_time = timezone.now()
            order.approval_comment = comment
            order.save()
            messages.success(request, f'出库单 "{order.record_no}" 已批准，共 {order.items.count()} 项物品已出库并扣减库存！')

        elif action == 'wh_reject':
            if order.status != '待仓管审批':
                messages.error(request, '该出库单状态不符')
                return redirect('stockout_detail', pk=order.pk)
            from inventory.permissions import has_permission
            if not has_permission(role, 'stockout', 'approve'):
                messages.error(request, '您没有仓管审批权限')
                return redirect('stockout_detail', pk=order.pk)
            # 驳回时释放锁定库存
            for item in order.items.select_related('supply').all():
                OfficeSupply.objects.filter(pk=item.supply.pk).update(
                    locked_quantity=F('locked_quantity') - item.quantity
                )
            order.status = '已拒绝'
            order.approver = request.user
            order.approval_time = timezone.now()
            order.approval_comment = comment
            order.save()
            messages.info(request, f'出库单 "{order.record_no}" 已被仓管驳回，已释放锁定库存')

        return redirect('stockout_detail', pk=order.pk)

    return redirect('stockout_detail', pk=order.pk)


# ==================== 出库管理 ====================
@login_required
def stockout_list(request):
    """出库记录查询"""
    query = request.GET.get('q', '')
    out_type = request.GET.get('out_type', '')
    status_filter = request.GET.get('status', '')

    records = get_visible_queryset(request.user, StockOutOrder).prefetch_related('items__supply').all()

    if query:
        records = records.filter(
            Q(record_no__icontains=query) | Q(recipient__icontains=query)
        )
    if out_type:
        records = records.filter(out_type=out_type)
    if status_filter:
        records = records.filter(status=status_filter)

    paginator = Paginator(records, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'inventory/stockout_list.html', {
        'page_obj': page_obj,
        'query': query,
        'out_type_filter': out_type,
        'status_filter': status_filter,
    })


@login_required
def stockout_create(request):
    """出库登记（支持多物品）——带库存锁定机制"""
    import json
    from django.db import transaction
    from django.db.models import F
    if request.method == 'POST':
        form = StockOutForm(request.POST, user=request.user)
        items_json = request.POST.get('items_json', '[]')
        try:
            items_data = json.loads(items_json)
        except (json.JSONDecodeError, TypeError):
            items_data = []

        if form.is_valid() and items_data:
            # 阶段1：合并相同物品并初步校验
            merged = {}
            errors = []
            error_indices = []
            for i, item in enumerate(items_data):
                supply_id = item.get('supply_id')
                qty = int(item.get('quantity', 0))
                try:
                    supply = OfficeSupply.objects.select_related('item_category').get(pk=supply_id)
                except OfficeSupply.DoesNotExist:
                    errors.append(f'物品 ID {supply_id} 不存在')
                    error_indices.append(i)
                    continue
                is_supply_disabled = ('停用' in (supply.status or '')) or (
                    supply.item_category is not None and not supply.item_category.is_active
                )
                if is_supply_disabled:
                    errors.append(f'{supply.code} - {supply.name} 已停用，禁止出库')
                    error_indices.append(i)
                elif qty <= 0:
                    errors.append(f'{supply.name}：出库数量必须大于 0')
                    error_indices.append(i)
                else:
                    sid = str(supply_id)
                    merged[sid] = merged.get(sid, 0) + qty

            if errors:
                for err in errors:
                    messages.error(request, err)
                restored = []
                for item in items_data:
                    supply_id = item.get('supply_id')
                    qty = int(item.get('quantity', 0))
                    try:
                        supply = OfficeSupply.objects.get(pk=supply_id)
                        restored.append({
                            'supply_id': supply_id,
                            'name': str(supply),
                            'specification': supply.specification or '',
                            'quantity': qty,
                            'stock': supply.available_quantity,
                            'unit': supply.unit or '',
                        })
                    except OfficeSupply.DoesNotExist:
                        restored.append({
                            'supply_id': supply_id,
                            'name': f'(不存在 ID:{supply_id})',
                            'specification': '',
                            'quantity': qty,
                            'stock': 0,
                            'unit': '',
                        })
                return render(request, 'inventory/stockout_form.html', {
                    'form': form, 'title': '出库登记',
                    'restored_items_json': json.dumps(restored, ensure_ascii=False),
                    'error_indices_json': json.dumps(error_indices),
                })

            # 阶段2：原子性校验可用库存并锁定（防止并发超卖）
            lock_errors = []
            try:
                with transaction.atomic():
                    for sid, qty in list(merged.items()):
                        supply = OfficeSupply.objects.select_for_update().get(pk=int(sid))
                        if qty > supply.available_quantity:
                            lock_errors.append(
                                f'{supply.name} 可用库存不足！可用：{supply.available_quantity} {supply.unit}，需要：{qty}'
                            )
                    if lock_errors:
                        raise ValueError('库存锁定失败')

                    # 全部校验通过，创建出库单并锁定库存
                    order = form.save(commit=False)
                    order.operator = request.user
                    order.status = '待审批'
                    order.save()

                    for sid, qty in merged.items():
                        supply = OfficeSupply.objects.get(pk=int(sid))
                        # 创建明细
                        StockOutItem.objects.create(
                            order=order, supply=supply, quantity=qty,
                            specification=supply.specification or '',
                            unit=supply.unit,
                            location=supply.location or '',
                            supplier=supply.supplier or '',
                        )
                        # 锁定库存（原子加法，使用update避免触发save中的F表达式比较）
                        OfficeSupply.objects.filter(pk=supply.pk).update(
                            locked_quantity=F('locked_quantity') + qty
                        )

            except ValueError:
                for err in lock_errors:
                    messages.error(request, err)
                return render(request, 'inventory/stockout_form.html', {
                    'form': form, 'title': '出库登记',
                    'restored_items_json': '[]',
                    'error_indices_json': '[]',
                })

            messages.success(request, f'出库单 "{order.record_no}" 已提交，等待审批！')
            return redirect('stockout_list')
        else:
            if not items_data:
                messages.error(request, '请至少添加一项出库物品')
    else:
        form = StockOutForm(user=request.user)
        # 快速出库：从入库单预填充物品
        from_stockin_id = request.GET.get('from_stockin')
        if from_stockin_id:
            try:
                stockin = StockInApplication.objects.prefetch_related('items__supply').get(pk=int(from_stockin_id))
                if stockin.status == '已批准':
                    restored = []
                    for item in stockin.items.all():
                        supply = item.supply
                        restored.append({
                            'supply_id': supply.id,
                            'name': supply.name,
                            'specification': supply.specification or '',
                            'quantity': item.quantity,
                            'stock': supply.available_quantity,
                            'unit': supply.unit or '',
                        })
                    return render(request, 'inventory/stockout_form.html', {
                        'form': form,
                        'title': '出库登记',
                        'restored_items_json': json.dumps(restored, ensure_ascii=False),
                        'error_indices_json': '[]',
                    })
            except (ValueError, StockInApplication.DoesNotExist):
                pass

    return render(request, 'inventory/stockout_form.html', {
        'form': form,
        'title': '出库登记',
        'restored_items_json': '[]',
        'error_indices_json': '[]',
    })


# ==================== 归还申请 ====================

@login_required
def stockout_detail(request, pk):
    """出库单明细查看"""
    order = get_object_or_404(
        StockOutOrder.objects.select_related('department', 'operator').prefetch_related('items__supply'),
        pk=pk
    )
    role = get_user_role(request.user)
    can_dept_approve = role in ('admin', 'dept_head') and order.status == '待审批'
    can_warehouse_approve = role in ('admin', 'warehouse') and order.status == '待仓管审批'
    # 查询关联的归还记录
    return_applications = order.returns.select_related('supply', 'department', 'operator', 'approver').order_by('-created_at')
    return render(request, 'inventory/stockout_detail.html', {
        'order': order,
        'can_dept_approve': can_dept_approve,
        'can_warehouse_approve': can_warehouse_approve,
        'return_applications': return_applications,
    })


@login_required
def stockout_edit(request, pk):
    """出库单编辑（仅待审批/待仓管审批/已拒绝状态，且仅操作员本人或管理员）"""
    order = get_object_or_404(StockOutOrder, pk=pk)
    role = get_user_role(request.user)
    if role != 'admin' and order.operator != request.user:
        messages.error(request, '您只能编辑自己创建的出库单')
        return redirect('stockout_list')
    if order.status not in ('待审批', '待仓管审批', '已拒绝'):
        messages.error(request, '只有待审批、待仓管审批或已拒绝的出库单才能修改！')
        return redirect('stockout_list')

    import json

    if request.method == 'POST':
        form = StockOutForm(request.POST, instance=order, user=request.user)
        items_json = request.POST.get('items_json', '[]')
        try:
            items_data = json.loads(items_json)
        except (json.JSONDecodeError, TypeError):
            items_data = []

        if form.is_valid() and items_data:
            from django.db import transaction
            from django.db.models import F

            # 阶段1：合并新明细
            merged_new = {}
            errors = []
            for item in items_data:
                supply_id = item.get('supply_id')
                qty = int(item.get('quantity', 0))
                try:
                    supply = OfficeSupply.objects.select_related('item_category').get(pk=supply_id)
                except OfficeSupply.DoesNotExist:
                    errors.append(f'物品 ID {supply_id} 不存在')
                    continue
                is_supply_disabled = ('停用' in (supply.status or '')) or (
                    supply.item_category is not None and not supply.item_category.is_active
                )
                if is_supply_disabled:
                    errors.append(f'{supply.code} - {supply.name} 已停用，禁止出库')
                elif qty <= 0:
                    errors.append(f'{supply.name}：出库数量必须大于 0')
                else:
                    sid = str(supply_id)
                    merged_new[sid] = merged_new.get(sid, 0) + qty

            if errors:
                for err in errors:
                    messages.error(request, err)
                return render(request, 'inventory/stockout_form.html', {
                    'form': form, 'title': '修改出库单',
                    'restored_items_json': '[]',
                    'error_indices_json': '[]',
                })

            # 阶段2：原子操作——释放旧锁定、校验新可用库存、加新锁定
            try:
                with transaction.atomic():
                    # 2a. 释放旧明细的锁定库存（使用update避免触发save中的F表达式比较）
                    old_items = list(order.items.select_related('supply').all())
                    for item in old_items:
                        OfficeSupply.objects.filter(pk=item.supply.pk).update(
                            locked_quantity=F('locked_quantity') - item.quantity
                        )

                    # 2b. 校验新明细的可用库存（注意：释放旧锁定后需要重新查询）
                    for sid, qty in merged_new.items():
                        supply = OfficeSupply.objects.select_for_update().get(pk=int(sid))
                        if qty > supply.available_quantity:
                            raise ValueError(
                                f'{supply.name} 可用库存不足！可用：{supply.available_quantity} {supply.unit}，需要：{qty}'
                            )

                    # 2c. 全部通过，保存表单并重建明细
                    form.save()
                    order.items.all().delete()
                    for sid, qty in merged_new.items():
                        supply = OfficeSupply.objects.get(pk=int(sid))
                        StockOutItem.objects.create(
                            order=order, supply=supply, quantity=qty,
                            specification=supply.specification or '',
                            unit=supply.unit,
                            location=supply.location or '',
                            supplier=supply.supplier or '',
                        )
                        # 锁定新库存
                        OfficeSupply.objects.filter(pk=supply.pk).update(
                            locked_quantity=F('locked_quantity') + qty
                        )

            except ValueError as e:
                # 回滚后需要重新锁定旧库存（因为事务已回滚）
                for item in old_items:
                    OfficeSupply.objects.filter(pk=item.supply.pk).update(
                        locked_quantity=F('locked_quantity') + item.quantity
                    )
                messages.error(request, str(e))
                return render(request, 'inventory/stockout_form.html', {
                    'form': form, 'title': '修改出库单',
                    'restored_items_json': '[]',
                    'error_indices_json': '[]',
                })

            messages.success(request, f'出库单 "{order.record_no}" 更新成功！')
            return redirect('stockout_list')
        else:
            if not items_data:
                messages.error(request, '请至少添加一项出库物品')
    else:
        form = StockOutForm(instance=order, user=request.user)

    # 预填充已有明细
    existing_items = []
    for item in order.items.select_related('supply').all():
        existing_items.append({
            'supply_id': item.supply.pk,
            'name': str(item.supply),
            'specification': item.supply.specification or '',
            'quantity': item.quantity,
            'stock': item.supply.available_quantity,
            'unit': item.supply.unit or '',
        })

    return render(request, 'inventory/stockout_form.html', {
        'form': form,
        'title': '修改出库单',
        'restored_items_json': json.dumps(existing_items, ensure_ascii=False),
        'error_indices_json': '[]',
    })


@login_required
def stockout_delete(request, pk):
    """出库单删除（仅待审批/已拒绝状态可删除，且仅操作员本人或管理员）"""
    from django.db.models import F
    order = get_object_or_404(StockOutOrder, pk=pk)
    role = get_user_role(request.user)
    if role != 'admin' and order.operator != request.user:
        messages.error(request, '您只能删除自己创建的出库单')
        return redirect('stockout_list')
    if order.status not in ('待审批', '已拒绝'):
        messages.error(request, '只有待审批或已拒绝的出库单才能删除！')
        return redirect('stockout_list')

    if request.method == 'POST':
        record_no = order.record_no
        # 释放锁定库存（待审批/已拒绝状态的出库单才有锁定）
        if order.status in ('待审批', '待仓管审批'):
            for item in order.items.select_related('supply').all():
                OfficeSupply.objects.filter(pk=item.supply.pk).update(
                    locked_quantity=F('locked_quantity') - item.quantity
                )
        order.delete()
        messages.success(request, f'出库单 "{record_no}" 已删除！')
        return redirect('stockout_list')
    return render(request, 'inventory/stockout_confirm_delete.html', {'order': order})



@login_required
def return_application_list(request):
    """归还申请列表"""
    returns = ReturnApplication.objects.select_related('supply', 'stockout_order', 'operator').order_by('-created_at')
    paginator = Paginator(returns, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    role = get_user_role(request.user)
    return render(request, 'inventory/return_list.html', {
        'page_obj': page_obj,
        'can_approve': role in ('admin', 'warehouse'),
    })


@login_required
def return_application_create(request):
    """归还申请（基于已审批出库单）"""
    # 获取已审批通过的出库单，并过滤掉所有物品均已完全归还的出库单
    approved_orders_all = StockOutOrder.objects.filter(
        status='已批准'
    ).select_related('department').prefetch_related('items__supply').order_by('-created_at')

    approved_orders = []
    for order in approved_orders_all:
        available_items = []
        for item in order.items.all():
            already_returned = ReturnApplication.objects.filter(
                stockout_order=order, supply=item.supply, status__in=['待审批', '已批准']
            ).aggregate(total=Sum('quantity'))['total'] or 0
            remaining = item.quantity - already_returned
            if remaining > 0:
                available_items.append({
                    'id': str(item.supply_id),
                    'name': item.supply.name,
                    'qty': item.quantity,
                    'remaining': remaining,
                })
        if available_items:
            order.available_items_json = json.dumps(available_items)
            approved_orders.append(order)

    if request.method == 'POST':
        order_id = request.POST.get('stockout_order')
        supply_id = request.POST.get('supply')
        quantity = int(request.POST.get('quantity', 0))
        returner = request.POST.get('returner', '').strip()
        department_id = request.POST.get('department', '')
        return_date = request.POST.get('return_date', '')
        reason = request.POST.get('reason', '')
        
        errors = []
        if not order_id:
            errors.append('请选择关联出库单')
        if not supply_id:
            errors.append('请选择归还物品')
        if quantity <= 0:
            errors.append('归还数量必须大于0')
        if not returner:
            errors.append('请填写归还人')
        if not return_date:
            errors.append('请选择归还日期')
        
        # 校验归还数量不超过出库数量
        if order_id and supply_id and quantity > 0:
            try:
                order = StockOutOrder.objects.get(pk=order_id)
                item = StockOutItem.objects.get(order=order, supply_id=supply_id)
                # 计算已归还数量
                already_returned = ReturnApplication.objects.filter(
                    stockout_order=order, supply_id=supply_id, status__in=['待审批', '已批准']
                ).aggregate(total=Sum('quantity'))['total'] or 0
                remaining = item.quantity - already_returned
                if quantity > remaining:
                    errors.append(f'归还数量不能超过可归还数量 {remaining}')
            except (StockOutOrder.DoesNotExist, StockOutItem.DoesNotExist):
                errors.append('出库单或物品不存在')
        
        if errors:
            for err in errors:
                messages.error(request, err)
        else:
            ret = ReturnApplication(
                stockout_order=order,
                supply_id=supply_id,
                quantity=quantity,
                returner=returner,
                department_id=department_id if department_id else None,
                return_date=return_date,
                reason=reason,
                operator=request.user,
                status='待审批',
            )
            supply = ret.supply
            ret.specification = supply.specification or ''
            ret.unit = supply.unit
            ret.location = supply.location or ''
            ret.supplier = supply.supplier or ''
            ret.save()
            
            messages.success(request, f'归还申请 "{ret.return_no}" 提交成功，待仓管审批！')
            return redirect('return_list')
    
    return render(request, 'inventory/return_form.html', {
        'approved_orders': approved_orders,
        'title': '归还申请',
    })


@login_required
def return_approval(request, pk):
    """仓管审批归还"""
    role = get_user_role(request.user)
    if role not in ('admin', 'warehouse'):
        messages.error(request, '您没有审批归还的权限')
        return redirect('return_list')
    
    ret = get_object_or_404(ReturnApplication, pk=pk)
    if ret.status != '待审批':
        messages.error(request, '该归还申请已审批')
        return redirect('return_list')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        comment = request.POST.get('comment', '')
        
        if action == 'approve':
            ret.status = '已批准'
            # 审批通过后增加库存
            supply = ret.supply
            supply.quantity += ret.quantity
            supply.save()
            messages.success(request, f'归还申请 "{ret.return_no}" 已批准，库存已更新！')
        elif action == 'reject':
            ret.status = '已拒绝'
            messages.info(request, f'归还申请 "{ret.return_no}" 已拒绝')
        
        ret.approver = request.user
        ret.approval_time = timezone.now()
        ret.approval_comment = comment
        ret.save()
        return redirect('return_list')
    
    return render(request, 'inventory/return_approval.html', {'ret': ret})


# ==================== IT设备管理 ====================
@login_required
def device_list(request):
    """计算机资源一览"""
    query = request.GET.get('q', '')
    category = request.GET.get('category', '')
    status = request.GET.get('status', '')
    
    devices = ITDevice.objects.select_related('device_type').all()
    
    if query:
        devices = devices.filter(
            Q(device_no__icontains=query) | Q(device_type__type_name__icontains=query)
        )
    if category:
        devices = devices.filter(device_type__category=category)
    if status:
        devices = devices.filter(status=status)
    
    paginator = Paginator(devices, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    categories = ComputerType.objects.values_list('category', flat=True).distinct()
    
    return render(request, 'inventory/device_list.html', {
        'page_obj': page_obj,
        'categories': categories,
        'query': query,
        'category_filter': category,
        'status_filter': status,
    })


@login_required
def device_create(request):
    """设备登记"""
    if request.method == 'POST':
        form = ITDeviceForm(request.POST)
        if form.is_valid():
            device = form.save()
            messages.success(request, f'设备 "{device.device_no}" 登记成功！')
            return redirect('device_list')
    else:
        form = ITDeviceForm()
    
    return render(request, 'inventory/device_form.html', {
        'form': form,
        'title': '设备登记',
        'action': '创建'
    })


@login_required
def device_update(request, pk):
    """设备信息变更"""
    device = get_object_or_404(ITDevice, pk=pk)
    if request.method == 'POST':
        form = ITDeviceForm(request.POST, instance=device)
        if form.is_valid():
            form.save()
            messages.success(request, '设备信息更新成功！')
            return redirect('device_list')
    else:
        form = ITDeviceForm(instance=device)
    
    return render(request, 'inventory/device_form.html', {
        'form': form,
        'device': device,
        'title': '设备变更',
        'action': '更新'
    })


@login_required
def device_delete(request, pk):
    """设备删除"""
    device = get_object_or_404(ITDevice, pk=pk)
    if request.method == 'POST':
        device.delete()
        messages.success(request, '设备记录已删除！')
        return redirect('device_list')
    return render(request, 'inventory/device_confirm_delete.html', {'device': device})


# ==================== 计算机类型管理 ====================
@login_required
def computer_type_list(request):
    """计算机类型一览"""
    types = ComputerType.objects.all().order_by('type_code')
    return render(request, 'inventory/computer_type_list.html', {'types': types})


@login_required
def computer_type_create(request):
    """类型追加"""
    if request.method == 'POST':
        form = ComputerTypeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '计算机类型添加成功！')
            return redirect('computer_type_list')
    else:
        form = ComputerTypeForm()
    
    return render(request, 'inventory/computer_type_form.html', {
        'form': form,
        'title': '类型追加',
        'action': '创建'
    })


@login_required
def computer_type_update(request, pk):
    """类型变更"""
    comp_type = get_object_or_404(ComputerType, pk=pk)
    if request.method == 'POST':
        form = ComputerTypeForm(request.POST, instance=comp_type)
        if form.is_valid():
            form.save()
            messages.success(request, '计算机类型更新成功！')
            return redirect('computer_type_list')
    else:
        form = ComputerTypeForm(instance=comp_type)
    
    return render(request, 'inventory/computer_type_form.html', {
        'form': form,
        'comp_type': comp_type,
        'title': '类型变更',
        'action': '更新'
    })


@login_required
def computer_type_delete(request, pk):
    """类型删除"""
    comp_type = get_object_or_404(ComputerType, pk=pk)
    if request.method == 'POST':
        comp_type.delete()
        messages.success(request, '计算机类型已删除！')
        return redirect('computer_type_list')
    return render(request, 'inventory/computer_type_confirm_delete.html', {'comp_type': comp_type})


# ==================== 统计报表 ====================
@login_required
def stockin_statistics(request):
    """入库统计"""
    from django.db.models.functions import TruncMonth
    
    # 获取日期范围参数
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    # 基础查询集
    queryset = StockInApplication.objects.filter(status='已批准')
    
    # 应用日期筛选
    if start_date:
        queryset = queryset.filter(approval_time__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(approval_time__date__lte=end_date)
    
    # 按月份统计（基于明细数量）
    monthly_stats = queryset.annotate(
        month=TruncMonth('approval_time')
    ).values('month').annotate(
        total_quantity=Sum('items__quantity'),
        count=Count('id')
    ).order_by('month')

    # 按类别统计（基于明细数量）
    category_stats = queryset.values(
        'items__supply__item_category__name'
    ).annotate(
        total_quantity=Sum('items__quantity'),
        count=Count('id')
    ).order_by('-total_quantity')

    # 计算汇总数据
    total_count = queryset.count()
    total_quantity = queryset.aggregate(total=Sum('items__quantity'))['total'] or 0
    
    context = {
        'monthly_stats': monthly_stats,
        'category_stats': category_stats,
        'start_date': start_date,
        'end_date': end_date,
        'total_count': total_count,
        'total_quantity': total_quantity,
    }
    return render(request, 'inventory/stockin_statistics.html', context)


@login_required
def stockout_statistics(request):
    """出库统计"""
    from django.db.models.functions import TruncMonth
    
    # 获取日期范围参数
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    # 基础查询集
    queryset = StockOutOrder.objects.all()

    # 应用日期筛选
    if start_date:
        queryset = queryset.filter(created_at__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(created_at__date__lte=end_date)

    # 按月份统计（基于明细数量）
    monthly_stats = StockOutItem.objects.filter(order__in=queryset).annotate(
        month=TruncMonth('order__created_at')
    ).values('month').annotate(
        total_quantity=Sum('quantity'),
        count=Count('id')
    ).order_by('month')

    # 按出库类型统计（基于明细数量）
    type_stats = StockOutItem.objects.filter(order__in=queryset).values('order__out_type').annotate(
        total_quantity=Sum('quantity'),
        count=Count('id')
    ).order_by('-total_quantity')

    # 计算汇总数据
    total_count = queryset.count()
    total_quantity = StockOutItem.objects.filter(order__in=queryset).aggregate(total=Sum('quantity'))['total'] or 0

    context = {
        'monthly_stats': monthly_stats,
        'type_stats': type_stats,
        'start_date': start_date,
        'end_date': end_date,
        'total_count': total_count,
        'total_quantity': total_quantity,
    }
    return render(request, 'inventory/stockout_statistics.html', context)


# ==================== Excel 导入导出 ====================
@login_required
def supply_export_excel(request):
    """办公用品库存表导出Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "办公用品库存"
    
    # 表头样式
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 表头
    headers = ['物品编码', '物品名称', '类别', '规格型号', '单位', '库存数量', '安全库存', 
               '存放位置', '供应商', '单价', '状态']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 数据
    supplies = OfficeSupply.objects.select_related('item_category').all()
    for row, supply in enumerate(supplies, 2):
        ws.cell(row=row, column=1, value=supply.code).border = border
        ws.cell(row=row, column=2, value=supply.name).border = border
        ws.cell(row=row, column=3, value=supply.item_category.name if supply.item_category else '').border = border
        ws.cell(row=row, column=4, value=supply.specification or '').border = border
        ws.cell(row=row, column=5, value=supply.unit).border = border
        ws.cell(row=row, column=6, value=supply.quantity).border = border
        ws.cell(row=row, column=7, value=supply.safety_stock).border = border
        ws.cell(row=row, column=8, value=supply.location or '').border = border
        ws.cell(row=row, column=9, value=supply.supplier or '').border = border
        ws.cell(row=row, column=10, value=float(supply.price)).border = border
        ws.cell(row=row, column=11, value=supply.status).border = border
    
    # 调整列宽
    column_widths = [15, 20, 15, 20, 10, 12, 12, 20, 20, 12, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # 生成响应
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="办公用品库存_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    return response


@login_required
def supply_import_excel(request):
    """办公用品库存表导入Excel"""
    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            update_existing = form.cleaned_data.get('update_existing', False)
            
            try:
                wb = load_workbook(excel_file)
                ws = wb.active
                
                imported_count = 0
                updated_count = 0
                error_count = 0
                errors = []
                
                # 从第2行开始读取（跳过表头）
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    try:
                        # 编码列保留兼容（可空），名称必填
                        raw_code = str(row[0] or '').strip()
                        name = str(row[1] or '').strip()
                        if not name:
                            continue

                        category_name = str(row[2] or '其他').strip()
                        specification = str(row[3] or '').strip()
                        unit = str(row[4] or '个').strip()
                        quantity = int(row[5] or 0)
                        safety_stock = int(row[6] or 10)
                        location = str(row[7] or '').strip()
                        supplier = str(row[8] or '').strip()
                        price = float(row[9] or 0)
                        status = str(row[10] or '正常').strip()

                        # 自动创建或获取类别
                        category, _ = ItemCategory.objects.get_or_create(
                            name=category_name,
                            parent=None,
                            defaults={
                                'code': f"CAT{ItemCategory.objects.count() + 1:03d}",
                                'sort_order': ItemCategory.objects.count(),
                                'is_active': True
                            }
                        )

                        # 更新逻辑：仅在勾选“更新已存在记录”且提供编码时按编码更新
                        existing = None
                        if update_existing and raw_code:
                            existing = OfficeSupply.objects.filter(code=raw_code).first()

                        if existing:
                            existing.name = name
                            existing.item_category = category
                            existing.specification = specification
                            existing.unit = unit
                            existing.quantity = quantity
                            existing.safety_stock = safety_stock
                            existing.location = location
                            existing.supplier = supplier
                            existing.price = price
                            existing.status = status
                            existing.save()
                            updated_count += 1
                        else:
                            # 创建新记录：编码由模型自动生成
                            OfficeSupply.objects.create(
                                name=name,
                                item_category=category,
                                specification=specification,
                                unit=unit,
                                quantity=quantity,
                                safety_stock=safety_stock,
                                location=location,
                                supplier=supplier,
                                price=price,
                                status=status
                            )
                            imported_count += 1
                        
                    except Exception as e:
                        error_count += 1
                        errors.append(f"第{row_idx}行: {str(e)}")
                
                # 显示结果
                if imported_count > 0:
                    messages.success(request, f'成功导入 {imported_count} 条记录！')
                if updated_count > 0:
                    messages.success(request, f'成功更新 {updated_count} 条记录！')
                if error_count > 0:
                    messages.warning(request, f'{error_count} 条记录导入失败。')
                    for err in errors[:5]:
                        messages.error(request, err)
                
                return redirect('supply_list')
                
            except Exception as e:
                messages.error(request, f'文件解析失败: {str(e)}')
                return redirect('supply_import_excel')
    else:
        form = ExcelImportForm()
    
    return render(request, 'inventory/supply_import.html', {'form': form, 'title': '导入办公用品'})


@login_required
def device_export_excel(request):
    """IT设备表导出Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "IT设备"
    
    # 表头样式
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 表头
    headers = ['设备编号', '设备类型', '资产编号', '序列号', '采购日期', '采购价格',
               '存放位置', '使用人', '使用部门', '状态', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 数据
    devices = ITDevice.objects.select_related('device_type').all()
    for row, device in enumerate(devices, 2):
        ws.cell(row=row, column=1, value=device.device_no).border = border
        ws.cell(row=row, column=2, value=device.device_type.type_name if device.device_type else '').border = border
        ws.cell(row=row, column=3, value=device.asset_no or '').border = border
        ws.cell(row=row, column=4, value=device.serial_no or '').border = border
        ws.cell(row=row, column=5, value=device.purchase_date.strftime('%Y-%m-%d') if device.purchase_date else '').border = border
        ws.cell(row=row, column=6, value=float(device.price)).border = border
        ws.cell(row=row, column=7, value=device.location or '').border = border
        ws.cell(row=row, column=8, value=device.user or '').border = border
        ws.cell(row=row, column=9, value=device.department or '').border = border
        ws.cell(row=row, column=10, value=device.status).border = border
        ws.cell(row=row, column=11, value=device.remarks or '').border = border
    
    # 调整列宽
    column_widths = [15, 20, 15, 20, 15, 12, 20, 15, 20, 10, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # 生成响应
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="IT设备_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    return response


@login_required
def device_import_excel(request):
    """IT设备表导入Excel"""
    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            update_existing = form.cleaned_data.get('update_existing', False)
            
            try:
                wb = load_workbook(excel_file)
                ws = wb.active
                
                imported_count = 0
                updated_count = 0
                error_count = 0
                errors = []
                
                # 从第2行开始读取（跳过表头）
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    try:
                        if not row[0] or not row[1]:  # 设备编号和设备类型必填
                            continue
                        
                        device_no = str(row[0]).strip()
                        type_name = str(row[1]).strip()
                        asset_no = str(row[2] or '').strip()
                        serial_no = str(row[3] or '').strip()
                        purchase_date = row[4]
                        price = float(row[5] or 0)
                        location = str(row[6] or '').strip()
                        user = str(row[7] or '').strip()
                        department = str(row[8] or '').strip()
                        status = str(row[9] or '库存').strip()
                        remarks = str(row[10] or '').strip()
                        
                        # 查找或创建设备类型
                        device_type, _ = ComputerType.objects.get_or_create(
                            type_name=type_name,
                            defaults={
                                'type_code': f'TYPE{ComputerType.objects.count() + 1:03d}',
                                'category': '主机',
                                'warranty_months': 36
                            }
                        )
                        
                        # 检查是否已存在
                        existing = ITDevice.objects.filter(device_no=device_no).first()
                        
                        if existing and update_existing:
                            # 更新现有记录
                            existing.device_type = device_type
                            existing.asset_no = asset_no
                            existing.serial_no = serial_no
                            if purchase_date and isinstance(purchase_date, str):
                                existing.purchase_date = purchase_date
                            existing.price = price
                            existing.location = location
                            existing.user = user
                            existing.department = department
                            existing.status = status
                            existing.remarks = remarks
                            existing.save()
                            updated_count += 1
                        elif not existing:
                            # 创建新记录
                            ITDevice.objects.create(
                                device_no=device_no,
                                device_type=device_type,
                                asset_no=asset_no,
                                serial_no=serial_no,
                                purchase_date=purchase_date if isinstance(purchase_date, str) else None,
                                price=price,
                                location=location,
                                user=user,
                                department=department,
                                status=status,
                                remarks=remarks
                            )
                            imported_count += 1
                        
                    except Exception as e:
                        error_count += 1
                        errors.append(f"第{row_idx}行: {str(e)}")
                
                # 显示结果
                if imported_count > 0:
                    messages.success(request, f'成功导入 {imported_count} 条记录！')
                if updated_count > 0:
                    messages.success(request, f'成功更新 {updated_count} 条记录！')
                if error_count > 0:
                    messages.warning(request, f'{error_count} 条记录导入失败。')
                    for err in errors[:5]:
                        messages.error(request, err)
                
                return redirect('device_list')
                
            except Exception as e:
                messages.error(request, f'文件解析失败: {str(e)}')
                return redirect('device_import_excel')
    else:
        form = ExcelImportForm()
    
    return render(request, 'inventory/device_import.html', {'form': form, 'title': '导入IT设备'})


@login_required
def supply_template_download(request):
    """下载办公用品导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "办公用品导入模板"
    
    # 表头样式
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 表头
    headers = ['物品编码', '物品名称', '类别', '规格型号', '单位', '库存数量', '安全库存', 
               '存放位置', '供应商', '单价', '状态']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 示例数据行（灰色提示）
    example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    example_font = Font(color="999999", italic=True)
    
    examples = [
        ['', 'A4打印纸', '耗材', '70g/500张', '包', 100, 20, 'A仓库', 'XX文具', 25.00, '正常'],
        ['', '中性笔', '文具', '0.5mm黑色', '支', 200, 50, 'A仓库', 'XX文具', 2.50, '正常'],
        ['', '订书机', '文具', '标准型', '个', 10, 5, 'A仓库', 'XX办公', 15.00, '正常'],
    ]
    
    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill
            cell.font = example_font
            cell.border = border
    
    # 调整列宽
    column_widths = [15, 20, 15, 20, 10, 12, 12, 20, 20, 12, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # 添加说明工作表
    ws_help = wb.create_sheet("填写说明")
    ws_help.column_dimensions['A'].width = 20
    ws_help.column_dimensions['B'].width = 60
    
    help_data = [
        ['字段名', '填写说明'],
        ['物品编码', '自动生成（分类编码-不重复数字），新建时留空；更新已有记录时可填现有编码'],
        ['物品名称', '必填，物品完整名称'],
        ['类别', '建议填写：文具、耗材、设备、其他'],
        ['规格型号', '可选，如：70g/500张、0.5mm黑色'],
        ['单位', '计量单位，如：个、包、盒、支'],
        ['库存数量', '数字，当前库存数量'],
        ['安全库存', '数字，低于此数量会预警，建议10-50'],
        ['存放位置', '可选，如：A仓库、B货架'],
        ['供应商', '可选，供应商名称'],
        ['单价', '数字，单位价格（元）'],
        ['状态', '可选，正常/低库存/停用，默认：正常'],
        ['', ''],
        ['注意事项', ''],
        ['1', '第1行为表头，请勿删除或修改'],
        ['2', '第2-4行为示例数据，导入前请删除'],
        ['3', '系统自动生成物品编码，页面内不可手动修改'],
        ['4', '所有数字字段请勿添加单位或特殊符号'],
    ]
    
    for row_idx, row_data in enumerate(help_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_help.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            if row_data[0] == '注意事项':
                cell.font = Font(bold=True, color="FF0000")
    
    # 生成响应
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="办公用品导入模板.xlsx"'
    return response


@login_required
def device_template_download(request):
    """下载IT设备导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "IT设备导入模板"
    
    # 表头样式
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 表头
    headers = ['设备编号', '设备类型', '资产编号', '序列号', '采购日期', '采购价格',
               '存放位置', '使用人', '使用部门', '状态', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 示例数据行（灰色提示）
    example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    example_font = Font(color="999999", italic=True)
    
    examples = [
        ['PC001', '台式机', 'ZC2024001', 'SN123456789', '2024-01-15', 5000.00, '机房A', '张三', '技术部', '使用中', '研发用'],
        ['PC002', '笔记本', 'ZC2024002', 'SN987654321', '2024-02-20', 8000.00, '办公室B', '李四', '销售部', '使用中', '出差用'],
        ['MN001', '显示器', 'ZC2024003', 'SN111222333', '2024-01-15', 1500.00, '机房A', '', '', '库存', '备用'],
    ]
    
    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill
            cell.font = example_font
            cell.border = border
    
    # 调整列宽
    column_widths = [15, 20, 15, 20, 15, 12, 20, 15, 20, 10, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # 添加说明工作表
    ws_help = wb.create_sheet("填写说明")
    ws_help.column_dimensions['A'].width = 20
    ws_help.column_dimensions['B'].width = 60
    
    help_data = [
        ['字段名', '填写说明'],
        ['设备编号', '必填，唯一标识，如：PC001、NB001'],
        ['设备类型', '必填，如：台式机、笔记本、显示器、服务器'],
        ['资产编号', '可选，公司资产编号'],
        ['序列号', '可选，设备序列号/SN码'],
        ['采购日期', '可选，格式：YYYY-MM-DD，如：2024-01-15'],
        ['采购价格', '数字，单位：元'],
        ['存放位置', '可选，如：机房A、办公室B、仓库'],
        ['使用人', '可选，当前使用人姓名'],
        ['使用部门', '可选，所属部门名称'],
        ['状态', '可选，库存/使用中/维修中/报废，默认：库存'],
        ['备注', '可选，其他说明信息'],
        ['', ''],
        ['注意事项', ''],
        ['1', '第1行为表头，请勿删除或修改'],
        ['2', '第2-4行为示例数据，导入前请删除'],
        ['3', '设备编号不能重复，重复时选择"更新已存在记录"可覆盖'],
        ['4', '设备类型不存在时会自动创建'],
        ['5', '日期格式必须为 YYYY-MM-DD，如：2024-01-15'],
    ]
    
    for row_idx, row_data in enumerate(help_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_help.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            if row_data[0] == '注意事项':
                cell.font = Font(bold=True, color="FF0000")
    
    # 生成响应
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="IT设备导入模板.xlsx"'
    return response


# ==================== 物品详情（属性编辑 + 操作记录）====================
@login_required
def supply_detail(request, pk):
    """物品详情页：属性编辑 + 操作记录"""
    supply = get_object_or_404(
        OfficeSupply.objects.select_related('item_category'),
        pk=pk
    )

    if request.method == 'POST':
        form = OfficeSupplyForm(request.POST, instance=supply)
        if form.is_valid():
            form.save()
            messages.success(request, f'物品 "{supply.name}" 更新成功！')
            return redirect('supply_detail', pk=supply.pk)
    else:
        form = OfficeSupplyForm(instance=supply)

    # 规格和单位始终可编辑（已移除库存限制）
    has_stock = False

    # 收集该物品的所有操作记录
    records = []

    # 入库记录
    stockin_items = StockInItem.objects.filter(
        supply=supply
    ).select_related(
        'application__applicant', 'application__department'
    ).order_by('-application__stockin_date')
    for item in stockin_items:
        records.append({
            'time': item.application.stockin_date,
            'type': '入库',
            'type_class': 'text-success',
            'no': item.application.application_no,
            'link_pk': item.application.pk,
            'link_url': 'stockin_application_detail',
            'quantity': item.quantity,
            'unit': item.unit or supply.unit,
            'specification': item.specification or supply.specification or '',
            'person': item.application.applicant.username if item.application.applicant else '-',
            'remark': item.application.reason or '',
        })

    # 出库记录
    stockout_items = StockOutItem.objects.filter(
        supply=supply
    ).select_related(
        'order__operator'
    ).order_by('-order__created_at')
    for item in stockout_items:
        records.append({
            'time': item.order.created_at.date(),
            'type': '出库',
            'type_class': 'text-danger',
            'no': item.order.record_no,
            'link_pk': item.order.pk,
            'link_url': 'stockout_detail',
            'quantity': -item.quantity,
            'unit': item.unit or supply.unit,
            'specification': item.specification or supply.specification or '',
            'person': item.order.recipient or '-',
            'remark': item.order.purpose or '',
        })

    # 归还记录
    returns = ReturnApplication.objects.filter(
        supply=supply
    ).select_related('department').order_by('-return_date')
    for ret in returns:
        records.append({
            'time': ret.return_date,
            'type': '归还',
            'type_class': 'text-primary',
            'no': ret.return_no,
            'quantity': ret.quantity,
            'unit': ret.unit or supply.unit,
            'specification': ret.specification or supply.specification or '',
            'person': ret.returner or '-',
            'remark': ret.reason or '',
        })

    # 按时间倒序
    records.sort(key=lambda r: str(r['time']) if r['time'] else '', reverse=True)

    return render(request, 'inventory/supply_detail.html', {
        'supply': supply,
        'form': form,
        'records': records,
        'has_stock': False,  # 已移除库存限制
    })


# ========== 注册 ==========

def register_view(request):
    """用户注册"""
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            from django.contrib.auth.models import User, Group
            user = User.objects.create_user(
                username=form.cleaned_data['username'],
                password=form.cleaned_data['password'],
                is_active=False,  # 待审核
            )
            profile = user.profile
            profile.name = form.cleaned_data['name']
            profile.phone = form.cleaned_data.get('phone', '')
            profile.department = form.cleaned_data['department']
            profile.applied_role = form.cleaned_data['applied_role']
            profile.is_pending = True
            profile.save()
            
            messages.success(request, '注册成功！请等待管理员审核通过后即可登录。')
            return redirect('login')
    else:
        form = RegisterForm()
    
    return render(request, 'inventory/register.html', {'form': form})


# ========== API ==========

def api_dept_head_check(request):
    """检查部门是否已有部门长"""
    dept_id = request.GET.get('dept_id')
    if not dept_id:
        return JsonResponse({'exists': False})
    
    try:
        dept = Department.objects.get(pk=dept_id)
    except Department.DoesNotExist:
        return JsonResponse({'exists': False})
    
    existing = check_dept_head_exists(dept)
    if existing:
        profile = getattr(existing, 'profile', None)
        name = profile.name if profile and profile.name else existing.username
        return JsonResponse({'exists': True, 'name': name, 'username': existing.username})
    return JsonResponse({'exists': False})


# ========== 用户管理 ==========

@role_required('admin')
def user_list(request):
    """用户列表"""
    from django.contrib.auth.models import User
    users = User.objects.select_related('profile', 'profile__department').all().order_by('-is_active', 'username')
    
    # 搜索
    q = request.GET.get('q', '')
    if q:
        users = users.filter(
            Q(username__icontains=q) |
            Q(profile__name__icontains=q) |
            Q(profile__phone__icontains=q)
        )
    
    # 角色筛选
    role_filter = request.GET.get('role', '')
    if role_filter:
        group_map = get_role_group_map()
        users = users.filter(groups__name=group_map.get(role_filter, ''))

    paginator = Paginator(users, 20)
    page = request.GET.get('page')
    users_page = paginator.get_page(page)

    return render(request, 'inventory/user_list.html', {
        'users': users_page,
        'q': q,
        'role_filter': role_filter,
        'role_choices': get_role_choices(),
    })


@role_required('admin')
def user_create(request):
    """管理员创建用户"""
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        name = request.POST.get('name', '').strip()
        phone = request.POST.get('phone', '').strip()
        dept_id = request.POST.get('department', '')
        role = request.POST.get('role', 'staff')
        
        from django.contrib.auth.models import User, Group
        
        if User.objects.filter(username=username).exists():
            messages.error(request, f'用户名 {username} 已存在')
            return redirect('user_create')
        
        # 仅非管理员/仓管员要求部门必填
        if role not in ('admin', 'warehouse') and not dept_id:
            messages.error(request, '所属部门为必填项')
            return redirect('user_create')
        
        # 部门长唯一性检查
        if role == 'dept_head' and dept_id:
            dept = Department.objects.get(pk=dept_id)
            existing = check_dept_head_exists(dept)
            if existing:
                p = getattr(existing, 'profile', None)
                n = p.name if p and p.name else existing.username
                messages.error(request, f'该部门已有部门长（{n}），请先更换角色再创建新部门长')
                return redirect('user_create')
        
        user = User.objects.create_user(username=username, password=password)
        profile = user.profile
        profile.name = name
        profile.phone = phone
        if dept_id:
            try:
                profile.department = Department.objects.get(pk=dept_id)
            except Department.DoesNotExist:
                pass
        profile.applied_role = role
        profile.save()
        
        # 分配角色组
        group_map = get_role_group_map()
        group_name = group_map.get(role, '普通用户')
        group = Group.objects.get_or_create(name=group_name)[0]
        user.groups.set([group])

        messages.success(request, f'用户 {username} 创建成功')
        return redirect('user_list')

    departments = Department.objects.filter(is_active=True).order_by('sort_order', 'code')
    return render(request, 'inventory/user_form.html', {
        'departments': departments,
        'role_choices': get_role_choices(),
        'action': '创建',
    })


@role_required('admin')
def user_edit(request, pk):
    """编辑用户"""
    from django.contrib.auth.models import User, Group
    user = get_object_or_404(User, pk=pk)
    profile = user.profile
    
    if request.method == 'POST':
        profile.name = request.POST.get('name', '').strip()
        profile.phone = request.POST.get('phone', '').strip()
        dept_id = request.POST.get('department', '')
        
        # 仅非管理员/仓管员要求部门必填
        if profile.role not in ('admin', 'warehouse') and not dept_id:
            messages.error(request, '所属部门为必填项')
            return redirect('user_edit', pk=pk)
        
        profile.department = Department.objects.get(pk=dept_id) if dept_id else None
        profile.save()
        
        user.is_active = 'is_active' in request.POST
        user.save()
        
        messages.success(request, f'用户 {user.username} 信息已更新')
        return redirect('user_list')
    
    departments = Department.objects.filter(is_active=True).order_by('sort_order', 'code')
    return render(request, 'inventory/user_form.html', {
        'edit_user': user,
        'profile': profile,
        'departments': departments,
        'role_choices': get_role_choices(),
        'action': '编辑',
    })


@role_required('admin')
def user_disable(request, pk):
    """停用/启用用户"""
    from django.contrib.auth.models import User
    user = get_object_or_404(User, pk=pk)
    
    if user == request.user:
        messages.error(request, '不能停用自己的账号')
        return redirect('user_list')
    
    # 停用时检查待处理数据
    if user.is_active:
        can_disable, pending_items = check_pending_before_role_change(user)
        if not can_disable:
            messages.error(request, f'无法停用用户 {user.username}，该用户有以下待处理数据：')
            for item in pending_items:
                messages.warning(request, item)
            return redirect('user_list')
    
    user.is_active = not user.is_active
    user.save()
    status = '启用' if user.is_active else '停用'
    messages.success(request, f'已{status}用户 {user.username}')
    return redirect('user_list')


@role_required('admin')
def user_password_reset(request, pk):
    """重置用户密码"""
    from django.contrib.auth.models import User
    user = get_object_or_404(User, pk=pk)
    
    if request.method == 'POST':
        form = PasswordResetForm(request.POST)
        if form.is_valid():
            user.set_password(form.cleaned_data['new_password'])
            user.save()
            messages.success(request, f'已重置用户 {user.username} 的密码')
            return redirect('user_list')
    else:
        form = PasswordResetForm()
    
    return render(request, 'inventory/user_password.html', {
        'form': form,
        'target_user': user,
    })


@role_required('admin')
def user_role_assign(request, pk):
    """分配角色"""
    from django.contrib.auth.models import User, Group
    user = get_object_or_404(User, pk=pk)
    
    if request.method == 'POST':
        new_role = request.POST.get('role', 'staff')
        new_dept_id = request.POST.get('department', '')
        
        # 检查待处理数据
        can_change, pending_items = check_pending_before_role_change(user)
        if not can_change:
            messages.error(request, f'无法变更角色，该用户有以下待处理数据：')
            for item in pending_items:
                messages.warning(request, item)
            return redirect('user_role_assign', pk=pk)
        
        # 如果分配部门长，检查唯一性
        if new_role == 'dept_head' and new_dept_id:
            dept = Department.objects.get(pk=new_dept_id)
            existing_head = check_dept_head_exists(dept, exclude_user=user)
            if existing_head:
                # 检查是否确认替换
                confirm = request.POST.get('confirm_replace')
                if confirm != '1':
                    profile_ex = getattr(existing_head, 'profile', None)
                    name_ex = profile_ex.name if profile_ex and profile_ex.name else existing_head.username
                    return render(request, 'inventory/user_role_assign.html', {
                        'target_user': user,
                        'existing_head': {'name': name_ex, 'username': existing_head.username},
                        'role_choices': get_role_choices(),
                        'departments': Department.objects.filter(is_active=True).order_by('sort_order', 'code'),
                        'new_role': new_role,
                        'new_dept_id': new_dept_id,
                    })
                else:
                    # 降级旧部门长前先检查其待处理数据
                    can_downgrade, pending_items = check_pending_before_role_change(existing_head)
                    if not can_downgrade:
                        messages.error(request, f'无法替换部门长，当前部门长 {existing_head.username} 有以下待处理数据：')
                        for item in pending_items:
                            messages.warning(request, item)
                        return redirect('user_role_assign', pk=pk)
                    # 降级旧部门长
                    group_map = get_role_group_map()
                    old_group = Group.objects.get(name=group_map.get('dept_head', '部门长'))
                    existing_head.groups.remove(old_group)
                    staff_group = Group.objects.get(name=group_map.get('staff', '普通用户'))
                    existing_head.groups.add(staff_group)

        # 更新部门
        profile = user.profile
        if new_dept_id:
            profile.department = Department.objects.get(pk=new_dept_id)
        profile.applied_role = new_role
        profile.save()

        # 更新角色组
        group_map = get_role_group_map()
        group_name = group_map.get(new_role, '普通用户')
        group = Group.objects.get_or_create(name=group_name)[0]
        user.groups.set([group])

        messages.success(request, f'已将用户 {user.username} 的角色设置为「{get_role_display_name(new_role)}」')
        return redirect('user_list')

    departments = Department.objects.filter(is_active=True).order_by('sort_order', 'code')
    return render(request, 'inventory/user_role_assign.html', {
        'target_user': user,
        'role_choices': get_role_choices(),
        'departments': departments,
    })


@role_required('admin')
def user_pending_list(request):
    """待审核用户列表"""
    from django.contrib.auth.models import User
    pending_users = User.objects.filter(
        profile__is_pending=True,
        is_active=False,
    ).select_related('profile', 'profile__department')
    
    return render(request, 'inventory/user_pending_list.html', {
        'pending_users': pending_users,
    })


@role_required('admin')
def user_approve(request, pk):
    """审核用户"""
    from django.contrib.auth.models import User, Group
    user = get_object_or_404(User, pk=pk)
    profile = user.profile
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'approve':
            # 可修改角色和部门
            role = request.POST.get('role', profile.applied_role)
            dept_id = request.POST.get('department', '')
            
            # 角色变更前检查待处理数据
            can_change, pending_items = check_pending_before_role_change(user)
            if not can_change:
                messages.error(request, f'无法变更角色，该用户有以下待处理数据：')
                for item in pending_items:
                    messages.warning(request, item)
                return redirect('user_approve', pk=pk)
            
            # 部门长唯一性检查
            if role == 'dept_head' and dept_id:
                dept = Department.objects.get(pk=dept_id)
                existing_head = check_dept_head_exists(dept, exclude_user=user)
                if existing_head:
                    confirm = request.POST.get('confirm_replace')
                    if confirm != '1':
                        profile_ex = getattr(existing_head, 'profile', None)
                        name_ex = profile_ex.name if profile_ex and profile_ex.name else existing_head.username
                        return render(request, 'inventory/user_approve.html', {
                            'target_user': user,
                            'existing_head': {'name': name_ex, 'username': existing_head.username},
                            'role_choices': get_role_choices(),
                            'departments': Department.objects.filter(is_active=True).order_by('sort_order', 'code'),
                            'selected_role': role,
                            'selected_dept': dept_id,
                        })
                    else:
                        # 降级旧部门长前先检查其待处理数据
                        can_downgrade, pending_items = check_pending_before_role_change(existing_head)
                        if not can_downgrade:
                            messages.error(request, f'无法替换部门长，当前部门长 {existing_head.username} 有以下待处理数据：')
                            for item in pending_items:
                                messages.warning(request, item)
                            return redirect('user_approve', pk=pk)
                        group_map = get_role_group_map()
                        old_group = Group.objects.get(name=group_map.get('dept_head', '部门长'))
                        existing_head.groups.remove(old_group)
                        staff_group = Group.objects.get(name=group_map.get('staff', '普通用户'))
                        existing_head.groups.add(staff_group)
            
            profile.applied_role = role
            if dept_id:
                profile.department = Department.objects.get(pk=dept_id)
            profile.is_pending = False
            profile.save()
            
            user.is_active = True
            user.save()
            
            # 分配角色组
            group_name = get_role_group_map().get(role, '普通用户')
            group = Group.objects.get_or_create(name=group_name)[0]
            user.groups.set([group])
            
            messages.success(request, f'已审核通过用户 {user.username}')
        
        elif action == 'reject':
            profile.is_pending = False
            profile.save()
            messages.info(request, f'已拒绝用户 {user.username} 的注册申请')
        
        return redirect('user_pending_list')
    
    departments = Department.objects.filter(is_active=True).order_by('sort_order', 'code')
    return render(request, 'inventory/user_approve.html', {
        'target_user': user,
        'role_choices': get_role_choices(),
        'departments': departments,
    })


@role_required('admin')
def permission_management(request):
    """权限管理总览（旧页面，保留兼容）"""
    from django.contrib.auth.models import User, Group
    from inventory.permissions import get_all_roles_permissions

    groups = Group.objects.all()
    role_data = []
    for role_key, role_name in get_role_choices():
        group = Group.objects.filter(name=role_name).first()
        users = group.user_set.select_related('profile', 'profile__department').all() if group else []
        role_data.append({
            'key': role_key,
            'name': role_name,
            'users': users,
            'count': users.count(),
        })

    return render(request, 'inventory/permissions.html', {
        'role_data': role_data,
    })


# ========== 角色与权限配置（可配置化） ==========

@role_required('admin')
def role_list(request):
    """角色列表与权限配置主页面"""
    from inventory.models import SystemRole, RolePermission
    from inventory.permissions import get_permission_matrix

    roles = SystemRole.objects.filter(is_active=True).order_by('sort_order', 'key')
    matrix = get_permission_matrix()

    scope_choices = [{'key': k, 'name': n} for k, n in RolePermission.SCOPE_CHOICES]

    return render(request, 'inventory/role_list.html', {
        'roles': roles,
        'matrix': matrix,
        'scope_choices': scope_choices,
    })


@role_required('admin')
def role_create(request):
    """创建新角色"""
    from django.contrib.auth.models import Group
    from inventory.models import SystemRole, RolePermission
    from inventory.permissions import BUILTIN_PERMISSIONS, clear_permission_cache

    if request.method == 'POST':
        key = request.POST.get('key', '').strip()
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        copy_from = request.POST.get('copy_from', '')

        # 校验
        if not key or not name:
            messages.error(request, '角色标识和名称不能为空')
            return redirect('role_create')
        if SystemRole.objects.filter(key=key).exists():
            messages.error(request, f'角色标识 "{key}" 已存在')
            return redirect('role_create')

        # 创建角色
        role = SystemRole.objects.create(
            key=key,
            name=name,
            description=description,
            is_builtin=False,
            is_active=True,
            sort_order=SystemRole.objects.count(),
        )

        # 同步创建 Django Group
        Group.objects.get_or_create(name=name)

        # 复制权限
        source_perms = BUILTIN_PERMISSIONS.get(copy_from, {})
        if not source_perms and copy_from:
            try:
                src_role = SystemRole.objects.get(key=copy_from)
                src_qs = RolePermission.objects.filter(role=src_role, is_enabled=True)
                source_perms = {}
                for p in src_qs:
                    source_perms.setdefault(p.module, {})[p.action] = p.scope
            except SystemRole.DoesNotExist:
                source_perms = BUILTIN_PERMISSIONS.get('staff', {})

        if not source_perms:
            source_perms = BUILTIN_PERMISSIONS.get('staff', {})

        for module, actions in source_perms.items():
            for action, scope in actions.items():
                RolePermission.objects.create(
                    role=role,
                    module=module,
                    action=action,
                    scope=scope,
                    is_enabled=True,
                )

        clear_permission_cache()
        messages.success(request, f'角色 "{name}" 创建成功')
        return redirect('role_list')

    # GET：展示可选的复制源角色
    builtin_roles = SystemRole.objects.filter(is_active=True).order_by('sort_order', 'key')
    return render(request, 'inventory/role_form.html', {
        'builtin_roles': builtin_roles,
        'action': '创建',
    })


@role_required('admin')
def role_edit(request, pk):
    """编辑角色信息（不修改权限，权限在 role_list 页面配置）"""
    from inventory.models import SystemRole
    from inventory.permissions import clear_permission_cache

    role = get_object_or_404(SystemRole, pk=pk)

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        is_active = 'is_active' in request.POST
        sort_order = request.POST.get('sort_order', '0').strip()

        if not name:
            messages.error(request, '角色名称不能为空')
            return redirect('role_edit', pk=pk)

        # 内置角色不允许修改标识和停用
        if role.is_builtin:
            is_active = True

        role.name = name
        role.description = description
        role.is_active = is_active
        role.sort_order = int(sort_order) if sort_order.isdigit() else 0
        role.save()

        # 同步更新 Django Group 名称
        from django.contrib.auth.models import Group
        old_group_name = get_role_group_map().get(role.key, name)
        group = Group.objects.filter(name=old_group_name).first()
        if group:
            group.name = name
            group.save()

        clear_permission_cache()
        messages.success(request, f'角色 "{name}" 已更新')
        return redirect('role_list')

    return render(request, 'inventory/role_form.html', {
        'role': role,
        'action': '编辑',
    })


@role_required('admin')
def role_delete(request, pk):
    """删除角色"""
    from inventory.models import SystemRole
    from inventory.permissions import clear_permission_cache

    role = get_object_or_404(SystemRole, pk=pk)

    if role.is_builtin:
        messages.error(request, '内置角色不能删除')
        return redirect('role_list')

    # 检查是否有用户关联此角色
    from django.contrib.auth.models import User
    group_name = get_role_group_map().get(role.key, role.name)
    group = User.groups.through.objects.filter(
        group__name=group_name
    ).exists()
    if group:
        messages.error(request, f'仍有用户属于角色 "{role.name}"，无法删除')
        return redirect('role_list')

    role.delete()
    clear_permission_cache()
    messages.success(request, f'角色 "{role.name}" 已删除')
    return redirect('role_list')


@role_required('admin')
def api_role_permissions(request, role_key):
    """API：获取某角色的权限配置"""
    from inventory.permissions import get_role_permissions
    from inventory.models import RolePermission

    perms = get_role_permissions(role_key)
    # 补充 action 和 module 的中文名称
    module_map = dict(RolePermission.MODULE_CHOICES)
    action_map = dict(RolePermission.ACTION_CHOICES)

    enriched = {}
    for module, actions in perms.items():
        enriched[module] = {
            'name': module_map.get(module, module),
            'actions': {
                action: {
                    'scope': scope,
                    'action_name': action_map.get(action, action),
                }
                for action, scope in actions.items()
            }
        }

    return JsonResponse({
        'role_key': role_key,
        'permissions': enriched,
    })


@csrf_exempt
@role_required('admin')
def api_save_permissions(request):
    """API：批量保存权限配置"""
    import json
    import logging
    from inventory.models import SystemRole, RolePermission
    from inventory.permissions import clear_permission_cache

    logger = logging.getLogger(__name__)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': '仅支持 POST 请求'}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError as e:
        logger.error(f'JSON 解析错误: {e}, body={request.body[:200]}')
        return JsonResponse({'success': False, 'error': f'无效的 JSON 数据: {e}'}, status=400)

    role_key = data.get('role_key')
    permissions = data.get('permissions', {})

    logger.info(f'保存权限请求: role_key={role_key}, permissions_keys={list(permissions.keys())}')

    if not role_key:
        return JsonResponse({'success': False, 'error': '缺少角色标识'}, status=400)

    if not permissions:
        return JsonResponse({'success': False, 'error': '权限数据为空，请至少修改一项权限'}, status=400)

    try:
        role = SystemRole.objects.get(key=role_key)
    except SystemRole.DoesNotExist:
        return JsonResponse({'success': False, 'error': f'角色 "{role_key}" 不存在'}, status=404)

    # 校验权限数据格式
    valid_modules = [m[0] for m in RolePermission.MODULE_CHOICES]
    valid_actions = [a[0] for a in RolePermission.ACTION_CHOICES]
    valid_scopes = [s[0] for s in RolePermission.SCOPE_CHOICES]

    errors = []
    for module, actions in permissions.items():
        if module not in valid_modules:
            errors.append(f'无效模块: {module}')
            continue
        if not isinstance(actions, dict):
            errors.append(f'模块 {module} 的操作数据格式错误')
            continue
        for action, scope in actions.items():
            if action not in valid_actions:
                errors.append(f'无效操作: {action}')
            if scope not in valid_scopes:
                errors.append(f'无效范围 "{scope}" (操作: {action})')

    if errors:
        logger.error(f'权限数据校验失败: {errors}')
        return JsonResponse({'success': False, 'error': '数据校验失败', 'details': errors}, status=400)

    try:
        updated = 0
        for module, actions in permissions.items():
            for action, scope in actions.items():
                is_enabled = scope != 'none'
                perm, created = RolePermission.objects.update_or_create(
                    role=role,
                    module=module,
                    action=action,
                    defaults={'scope': scope, 'is_enabled': is_enabled}
                )
                updated += 1

        clear_permission_cache()
        logger.info(f'权限保存成功: role={role_key}, updated={updated}')

        return JsonResponse({
            'success': True,
            'message': f'已更新 {updated} 条权限配置',
            'role_key': role_key,
        })
    except Exception as e:
        logger.exception(f'权限保存异常: role={role_key}')
        return JsonResponse({'success': False, 'error': f'保存失败: {str(e)}'}, status=500)


# ========== 个人中心 ==========

@login_required
def profile_view(request):
    """个人信息"""
    profile = request.user.profile
    if request.method == 'POST':
        form = ProfileForm(request.POST)
        if form.is_valid():
            profile.name = form.cleaned_data['name']
            profile.phone = form.cleaned_data['phone']
            profile.save()
            messages.success(request, '个人信息已更新')
            return redirect('profile')
    else:
        form = ProfileForm(initial={
            'name': profile.name,
            'phone': profile.phone,
        })
    
    return render(request, 'inventory/profile.html', {
        'form': form,
        'profile': profile,
    })


@login_required
def profile_password(request):
    """修改密码"""
    if request.method == 'POST':
        form = PasswordResetForm(request.POST)
        if form.is_valid():
            request.user.set_password(form.cleaned_data['new_password'])
            request.user.save()
            messages.success(request, '密码已修改，请重新登录')
            return redirect('login')
    else:
        form = PasswordResetForm()
    
    return render(request, 'inventory/profile_password.html', {'form': form})
